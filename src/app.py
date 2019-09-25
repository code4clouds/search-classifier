# Import required modules.
from azure.cognitiveservices.search.websearch import WebSearchAPI
from azure.cognitiveservices.search.websearch.models import SafeSearch
from msrest.authentication import CognitiveServicesCredentials
from openpyxl import load_workbook
import os


# Make a request. Replace Yosemite if you'd like.
def bingSearch(client, categories, searchTerm): 
    web_data = client.web.search(query=searchTerm)

    '''
    Web pages
    If the search response contains web pages, the first result's name and url
    are printed.
    '''
    category_detected = {}
    if hasattr(web_data.web_pages, 'value'):
        print("\r\nSearched for {} returned {} results".format( searchTerm, len(web_data.web_pages.value)))
        
        for web_page in web_data.web_pages.value:
            # print("First web page name: {} ".format(web_page.name))
            # print("First web page URL: {} ".format(web_page.url))
            # print("First web page description: {} ".format(web_page.snippet))    
            for key in categories:
                for cat_term in categories[key]:
                    if cat_term in web_page.snippet.lower():
                        if key in category_detected:
                            category_detected[key] += 1
                        else :
                            category_detected[key] = 1
        for key_category in category_detected:
            print("Category: {} {} ".format(key_category, category_detected[key_category]))    
    else:
        print("Didn't find any web pages...")

    return category_detected


def main():   
    # Load environmental variables
    SC_ENDPOINT = os.environ['SC_ENDPOINT']
    SC_KEY = os.environ['SC_KEY']

    # Instantiate the client and replace with your endpoint.
    client = WebSearchAPI(CognitiveServicesCredentials(SC_KEY), base_url = SC_ENDPOINT)

    categories = { 'data': [ 'data', 'operating system', 'container', 'dell', 'emc', 'networking', 'virtualization', 'db', 'database', 'relational', 'erp'],
                    'ml' : [ 'machine learning', 'ml', 'intelligence', 'artificial', 'science'],
                    'storage': ['storage', 'nas'],
                    'iot': ['iot', 'thing', 'internet of things','thermostat','bluetooth','modem', 'hardware', 'camera'],
                    'devops': ['devops', 'code', 'pipeline', 'api', 'kubernetes', 'k8s', 'gaming'],
                    'securty': ['cissp', 'hacker', 'security','defense', 'keys', 'privacy', 'virus', 'anti-virus'],
                    'health': ['hipaa'],
                    'media': ['streaming', 'video']
                    }

    # Load worksheet
    wb = load_workbook(filename = 'kusto_output.xlsx')
    ws = wb.active
    colA = ws['A']

    # Add Column for category in the sheet
    ws.insert_cols(2,2)
    ws['B1'] = 'Categories'
    ws['C1'] = 'Categories'

    # Search the tem and add it to the column starting at row B2
    for index, row in enumerate(colA[1:14176], start = 2):
        categories_found = bingSearch(client, categories, row.value)
        category_csv = ''
        for category in sorted(categories_found, key=categories_found.get):
            category_csv = category_csv + ',' + category
        ws['C' + str(index)] = category_csv
        ws['B' + str(index)] = category
        print("MAX: " + category)
        index += 1
    
    wb.save(filename = 'output.xlsx')

if __name__ == "__main__":
    main()

